import openpyxl as xls
import calc_eleven as calc
import sin_def

# workbook as wb
# worksheet as ws


# 单元格属性
cell_heigh = 54
cell_width = 25
cell_font = xls.styles.Font(bold=True, size=20)
cell_border = xls.styles.Border(bottom=xls.styles.Side(
    border_style="dashed", color='000000'))
cell_align = xls.styles.Alignment(horizontal='left', vertical='center')
# 页边距
page_margins = xls.worksheet.page.PageMargins(
    top=0.8, bottom=0.0, left=0.2, right=0.2)
# 每天题目数量
day_calc = 20
# 每页可做几天
day_nums = 3
# 每页题目总数
all_calc = day_calc * day_nums
# 每页行数
calc_row = 15
# 每页列数
calc_col = int(all_calc / calc_row)
# 一次包括几周题目
week_calc = 4


def xls_calc():
    wb = xls.Workbook()
    ws = wb.active
    ws.title = 'eleven'
    # print(ws.title)

    for i in range(0, week_calc*2):
        eleven = calc.c_calc(all_calc, 10)
        eleven.create()
        list_ques = eleven.get_question()
        list_calc_ceil = sin_def.list_ceil(list_ques, calc_col)
        for que in list_calc_ceil:
            ws.append(que)

    rows = ws.max_row
    cols = ws.max_column
    for r in range(1, rows+1):
        for c in range(1, cols+1):
            ws.cell(r, c).font = cell_font
            ws.cell(r, c).alignment = cell_align
            if (r % (calc_row/day_nums)) == 0:
                ws.cell(r, c).border = cell_border

    for r in range(1, rows+1):
        ws.row_dimensions[r].height = cell_heigh

    for c in range(1, cols+1):
        ws.column_dimensions[xls.utils.get_column_letter(c)].width = cell_width

    ws.page_margins = page_margins

    wb.save(sin_def.get_file_addr('calc', 'xlsx'))


if __name__ == '__main__':
    xls_calc()
