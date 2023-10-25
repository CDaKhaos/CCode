import openpyxl as xls
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

wb = xls.load_workbook('Data.xlsx')
sheet = wb['Sheet1']

max_row = sheet.max_row
max_col = sheet.max_column

# print(max_row, max_col)

def read_col(col_name):
    list_value: list = []
    col = sheet[col_name]
    for i in range(1, 20):
        list_value.append(col[i].value)
    return list_value
    

list_lon: list = read_col('L')
list_lat: list = read_col('M')

lon_ave = (max(list_lon) + min(list_lon)) / 2
lat_ave = (max(list_lat) + min(list_lat)) / 2

nX = 500

# print((max(list_lon) - lon_ave) * nX) 
# print((min(list_lon) - lon_ave) * nX) 

# print((max(list_lat) - lat_ave) * nX) 
# print((min(list_lat) - lat_ave) * nX) 

# print('########')

#####################################################
col_to_read = [0, 5, 7, 11, 12]

# 读取1到300行的数据  
list_data:list = []
list_lon:list = []
list_lat:list = []
list_time:list = []
list_doa:list = []
list_freq:list = []
for row in sheet.iter_rows(min_row=2, max_row=292, min_col=1, max_col=15, values_only=True):  
    list_freq.append(row[0])
    time = row[5].timestamp()
    list_time.append(time)
    doa = row[7] / 10
    list_doa.append(doa)
    lon = row[11]
    list_lon.append(lon)
    lat = row[12]
    list_lat.append(lat)
    list_data.append([doa, lon, lat])   
# print(list_data)

def draw_plt_scatter(row, col, index, list_data:list, title):
    plt.subplot(row, col,index)
    plt.scatter(list_time, list_data)
    plt.title(title)

def draw_plt_polt(col, index, list_x, list_y, title):
    plt.subplot(2, col, index)
    plt.plot(list_x, list_y, color="#FF3B1D", marker='*')
    plt.title(title)

def draw_ori():
    plt.figure()
    draw_plt_scatter(2, 2, 1, list_freq, "freq")
    draw_plt_scatter(2, 2, 2, list_doa,  "doa")
    draw_plt_scatter(2, 2, 3, list_lon, "lon")
    draw_plt_scatter(2, 2, 4, list_lat, "lat")
    plt.legend()
    plt.show()

draw_ori()

#############################
"""
data = np.stack((list_time, list_freq), axis=1)
# print(data)

# KMeans
from sklearn.cluster import KMeans
y_pred = KMeans().fit_predict(data)
plt.subplot(1,2,1)
plt.scatter(data[:, 0], data[:, 1], c=y_pred)
plt.title("KMeans")


# DBSCAN
from sklearn.cluster import DBSCAN
dbscan = DBSCAN(eps=3, min_samples=10).fit_predict(data)
plt.subplot(1,2,2)
plt.scatter(data[:, 0], data[:, 1], c=dbscan)
plt.title("DBSCAN")

plt.show()
"""
# print(clusters)

